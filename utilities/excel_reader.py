from openpyxl import load_workbook

def get_login_data():
    wb = load_workbook("test_data/login_data.xlsx")
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data